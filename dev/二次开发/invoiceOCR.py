import fitz
# pip install potencent
from poocr.api.ocr import VatInvoiceOCR
from pathlib import Path
# pip install openpyxl
import openpyxl
from openpyxl.styles import Font, Alignment, colors
from openpyxl.utils import get_column_letter
import json
from datetime import datetime
# pip install PySimpleGUI
import PySimpleGUI as sg


class InvoiceInfo(object):
    def __init__(self, dir_path: str, output: str):
        """
        :param dir_path: 输入需要识别的文件夹路径
        :param output: 输出excel文件路径
        """
        # 防止空值运行
        self.DirPath = dir_path.strip()
        # 保存发票数据的列表以及标题行(需要项目增加修改此处)
        self. invoice_list = [['销售方名称', '发票号码', '开票日期', '货物或应税劳务、服务名称', '税率', '发票金额', '发票税额', '价税合计', '发票类型', '是否盖章']]
        # 创建文件路径对象
        self.dir_path = Path(self.DirPath)
        # 保存需要删除的图片路径
        self.del_pic = []
        # 保存需要删除文件夹路径
        self.del_dir = []
        # 存放所有需识别文件的路径
        self.file_path_list = []
        # 输出excel文件路径
        self.output = output

    # 获取发票路径
    def all_path(self):
        # 遍历出所有的图片文件(图片文件格式只支持jpg, jpeg, png格式)
        for suffix in ["*.jpg", "*.png", "*.jpeg"]:
            # 获取指定格式的列表
            file_path = list(self.dir_path.rglob(suffix))
            # 并将列表添加到存放所有路径的列表中
            self.file_path_list.append(file_path)

        # 将所有PDF转化为图片
        # 获取所有PDF格式的路径的列表
        pdf_all_path = list(self.dir_path.rglob("*.pdf"))
        # 遍历PDF路径列表
        for pdf in pdf_all_path:
            # 打开PDF文件
            pdf_file = fitz.open(pdf)
            # 新建文件夹路径
            pdf_name = f'{pdf}'.replace(".pdf", "")
            # 创建与pdf名称相同的文件夹(用于保存图片)
            # 文件夹路径不存在创建文件夹
            if not Path(pdf_name).exists():
                Path(pdf_name).mkdir()
                # 并将文件夹添加到删除列表便于最后删除
                self.del_dir.append(pdf_name)
                sg.cprint("\r文件路径:", pdf, end=";")
                try:
                    # 遍历PDF的每一页
                    for i, page in enumerate(pdf_file):
                        sg.cprint("\r转换为图片进度:", i + 1, "/", pdf_file.page_count, end=";")
                        sg.cprint("")

                        # 旋转角度
                        rotate = int(0)
                        # 缩放系数为3
                        zoom_x = 3.0
                        zoom_y = 3.0

                        # 定义图片
                        trans = fitz.Matrix(zoom_x, zoom_y).prerotate(rotate)
                        pm = page.get_pixmap(matrix=trans, alpha=False)
                        # 保存图片
                        root = f"{pdf_name}/{pdf.name.replace('.pdf','')}_{str(i+1).zfill(4)}.png"
                        pm.save(root)
                        # 添加到删除图片的列表中,便于最后删除
                        self.del_pic.append(root)
                except Exception as e:
                    # 返回
                    sg.cprint(f'错误信息{e}', f'异常文件{pdf}')
                    continue
        # 将PDF转换的图片路径添加file_path_list中
        self.file_path_list.append(self.del_pic)
        # print(self.file_path_list)

    # 提取发票信息1
    @staticmethod
    def read_invoice1(info_list: list, value: list, information: str):
        """
        :param info_list: 用于存储该张发票的信息的列表
        :param value: 提取的数据列表
        :param information: 发票项目名称
        :return:
        """
        # 创建一个列表
        invoice_info = []
        # 遍历列表
        for j in range(0, len(value)):
            # 如果key等于发票项目名称
            if value[j]["Name"] == information:
                # 将字典的值添加到创建的列表中
                invoice_info.append(value[j]['Value'])
        # 如果列表存在
        if len(invoice_info) > 0:
            # 去重列表数据
            invoice_info = list(set(invoice_info))
            # 如果列表中只有一个值
            if len(invoice_info) == 1:
                # 将值添加到列表中
                info_list.append(invoice_info[0])
            else:
                # 列表的值大于一个用逗号连接
                info_list.append(",".join(invoice_info))
        else:
            # 不存在插入空值
            info_list.append("")

    # 提取发票信息2
    @staticmethod
    def read_invoice2(info_list: list, value: list, information: str, replace_value: str):
        """
        :param info_list: 用于存储该张发票的信息的列表
        :param value: 提取的数据列表
        :param information: 发票项目名称
        :param replace_value: 需要替换的值
        :return:
        """
        # 创建一个列表
        invoice_info = []
        # 遍历列表
        for j in range(0, len(value)):
            # 如果key等于发票项目名称
            if value[j]["Name"] == information:
                # 将字典的值添加到创建的列表中
                invoice_date = value[j]['Value']
                # 如果值存在替换的内容
                if invoice_date.find(replace_value) != -1:
                    # 去除替换的值
                    invoice_date = invoice_date.replace(replace_value, '')
                    # 添加到创建的列表中
                    invoice_info.append(invoice_date)
                else:
                    # 否则添加到创建的列表中
                    invoice_info.append(invoice_date)
        # 如果列表存在
        if len(invoice_info) > 0:
            # 去重列表数据
            invoice_info = list(set(invoice_info))
            # 如果列表中只有一个值
            if len(invoice_info) == 1:
                # 将值添加到列表中
                info_list.append(invoice_info[0])
            else:
                # 列表的值大于一个用逗号连接
                info_list.append(",".join(invoice_info))
        else:
            # 不存在插入空值
            info_list.append("")

    # 提取是否盖章信息
    @staticmethod
    def read_invoice_seal(info_list: list, value: list, information: str):
        """
        :param info_list: 用于存储该张发票的信息的列表
        :param value: 提取的数据列表
        :param information: 发票项目名称
        :return:
        """
        # 创建一个列表
        invoice_info = []
        # 遍历列表
        for j in range(0, len(value)):
            # 如果key等于发票项目名称
            if value[j]["Name"] == information:
                # 获取印章的值
                seal = value[j]['Value']
                # 如果值等于"1"
                if seal == '1':
                    # 等于"True"
                    invoice_seal = 'True'
                else:
                    # 否则等于"False"
                    invoice_seal = 'False'
                # 将值添加到创建的列表中
                invoice_info.append(invoice_seal)
        # 如果列表存在
        if len(invoice_info) > 0:
            # 去重列表数据
            invoice_info = list(set(invoice_info))
            # 如果列表中只有一个值
            if len(invoice_info) == 1:
                # 将值添加到列表中
                info_list.append(invoice_info[0])
            else:
                # 列表的值大于一个用逗号连接
                info_list.append(",".join(invoice_info))
        else:
            # 不存在插入空值
            info_list.append("")

    # 提取日期
    @staticmethod
    def read_invoice_date(info_list: list, value: list, information: str):
        """
        :param info_list: 用于存储该张发票的信息的列表
        :param value: 提取的数据列表
        :param information: 发票项目名称
        :return:
        """
        # 创建一个列表
        invoice_info = []
        # 遍历列表
        for j in range(0, len(value)):
            # 如果key等于发票项目名称
            if value[j]["Name"] == information:
                # 将字典的值添加到创建的列表中
                invoice_date = value[j]['Value']
                # 如果字典的值中同时存在年月日
                if invoice_date.find('年') != -1 and invoice_date.find('月') != -1 and invoice_date.find('日') != -1:
                    # 将值替换为 例:2022年2月4日 → 2022/2/4
                    invoice_date = invoice_date.replace('年', '/').replace('月', '/').replace('日', '')
                    # 将值添加到创建的列表中
                    invoice_info.append(invoice_date)
                else:
                    # 将值添加到创建的列表中
                    invoice_info.append(invoice_date)
        # 如果列表存在
        if len(invoice_info) > 0:
            # 去重列表数据
            invoice_info = list(set(invoice_info))
            # 如果列表中只有一个值
            if len(invoice_info) == 1:
                # 将值添加到列表中
                info_list.append(invoice_info[0])
            else:
                # 列表的值大于一个用逗号连接
                info_list.append(",".join(invoice_info))
        else:
            # 不存在插入空值
            info_list.append("")

    # 识别发票信息
    def read_info(self, file_path: str):
        """
        :param file_path: 发票文件路径
        :return:
        """
        # 识别发票文件
        res = VatInvoiceOCR(img_path=file_path)
        if res:
            # 将识别内容中的null替换为"null"
            res_str = f"{res}".replace('null', '"null"')
            # 将字符串转换为字典
            res_dic = json.loads(res_str)
            # 获取字典中'VatInvoiceInfos'的值
            invoice_value = res_dic['VatInvoiceInfos']
            print(invoice_value)
            # 创建一个存储发票的列表
            info_list = []

            # 提取发票信息
            self.read_invoice1(info_list, invoice_value, "销售方名称")
            self.read_invoice2(info_list, invoice_value, "发票号码", 'No')
            self.read_invoice_date(info_list, invoice_value, "开票日期")
            self.read_invoice1(info_list, invoice_value, "货物或应税劳务、服务名称")
            self.read_invoice1(info_list, invoice_value, "税率")
            self.read_invoice2(info_list, invoice_value, "合计金额", '¥')
            self.read_invoice2(info_list, invoice_value, "合计税额", '¥')
            self.read_invoice2(info_list, invoice_value, "小写金额", '¥')
            self.read_invoice1(info_list, invoice_value, "发票类型")
            self.read_invoice_seal(info_list, invoice_value, "是否有公司印章")

            # 将提取发票信息添加到列表
            self.invoice_list.append(info_list)
            sg.cprint(f"{file_path}发票提取成功")
        else:
            sg.cprint(f"{file_path}发票提取失败")

    # 写入excel
    def write_to_excel(self):
        # 创建一个工作簿
        wb = openpyxl.Workbook()
        # 打开当前工作表
        sh = wb.active
        # 把工作表名修改为'发票信息'
        sh.title = '发票信息'
        # 遍历发票信息列表
        for invoice in self.invoice_list:
            # 将发票信息添加到工作表中
            sh.append(invoice)

        # 标题字体（楷体，13号字，加粗，黑色）
        font1 = Font(name='楷体', size=13, bold=True, color=colors.BLACK)
        # 数据字体（楷体，11号字，黑色）
        font2 = Font(name='楷体', size=11, color=colors.BLACK)

        # 设置标题行格式
        for cells1 in sh.iter_rows(min_row=1, max_row=1):
            for cell1 in cells1:
                # 将第一行标题居中,添加font1字体
                cell1.font = font1
                cell1.alignment = Alignment(horizontal='centerContinuous', vertical='center')

        # 设置数据行格式
        # 全部列数据居中
        for cells4 in sh.iter_rows(min_row=2):
            for cell4 in cells4:
                # 将所有数据居中,添加font2字体
                cell4.font = font2
                cell4.alignment = Alignment(horizontal='centerContinuous', vertical='center')

        # A列数据靠左
        for cells2 in sh.iter_rows(min_row=2, max_col=1):
            for cell2 in cells2:
                cell2.alignment = Alignment(horizontal='left', vertical='center')

        # D列数据靠左
        for cells3 in sh.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell3 in cells3:
                cell3.alignment = Alignment(horizontal='left', vertical='center')

        # 设置宽度
        width_list = [25, 15, 15, 30, 15, 15, 15, 15, 17, 15]
        for i in range(0, 10):
            sh.column_dimensions[get_column_letter(i + 1)].width = width_list[i]

        # 保存Excel文件
        dt = datetime.now()
        wb.save(f"{self.output}/{str(dt.year).zfill(4)}{str(dt.month).zfill(2)}"
                f"{str(dt.day).zfill(2)}{str(dt.hour).zfill(2)}{str(dt.minute).zfill(2)}"
                f"{str(dt.second).zfill(2)}发票信息.xlsx")
        # 关闭工作簿
        wb.close()

    # 删除图片和文件夹
    def delete_pic(self):
        # 遍历列表删除创建的图片
        for pdf in self.del_pic:
            Path(pdf).unlink()
        # 遍历列表删除创建的文件夹
        for dir in self.del_dir:
            Path(dir).rmdir()

    # 提取发票信息到Excel
    def invoice_info(self):
        # 获取文件路径
        self.all_path()
        # 遍历列表
        for files in self.file_path_list:
            for file in files:
                try:
                    # 识别发票
                    self.read_info(file)
                except Exception as err:
                    sg.cprint(file, f"错误信息：{err}", sep="\n", text_color="red")
                    continue
        # 写入Excel
        self.write_to_excel()
        # 删除图片和文件夹
        self.delete_pic()

    # 优化
    def InvoiceInfo(self):
        try:
            if self.DirPath and self.output:
                # 调用合并函数
                self.invoice_info()
                sg.cprint("=" * 12, 'END', '=' * 12)
            else:
                sg.cprint("参数不能为空,请补齐参数!!!")
        except Exception as e:
            sg.cprint(f"错误信息：{e}", sep="\n", text_color="red")



# if __name__ == '__main__':
    # InvoiceInfo(r'发票位置的文件夹', r'输出路径').InvoiceInfo()
